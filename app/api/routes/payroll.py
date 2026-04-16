# app/api/routes/payroll.py
"""
Payroll routes.

Primary endpoints
─────────────────
  GET /payroll/monthly-report/{uid}?year=2026&month=4
      Single-employee monthly report in the standard format.
      Add &export=csv or &export=excel for file download.

  GET /payroll/monthly-report-all?year=2026&month=4
      All active employees, same format, JSON response.
      Add &export=csv or &export=excel for file download.

Row format (matches the requested sample):
  Sno | ID | Employee Name | Date | InTime-1 | OutTime-1 |
  InTime-2 | OutTime-2 | Shift | Total Duration | Status | Remarks

Multi-month range:
  GET /payroll/range-report/{uid}?start_date=2026-01-01&end_date=2026-04-30&export=excel
  Returns one sheet per month in the Excel file.
"""

import json
import calendar
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.response import success_response, error_response
from app.models.attendance import ProcessedAttendance, AttendanceStatus
from app.models.user import User
from app.config import get_settings

settings = get_settings()
router   = APIRouter(prefix="/payroll", tags=["Payroll & Reports"])

# ── formatting helpers ────────────────────────────────────────────────────── #

def _fmt_time(iso_str: Optional[str]) -> str:
    """'2026-04-04T10:00:00' → '10:00 AM'"""
    if not iso_str:
        return "-"
    try:
        return datetime.fromisoformat(iso_str).strftime("%-I:%M %p")
    except Exception:
        return "-"


def _fmt_hours(h: float) -> str:
    """9.0 → '9.00 hrs'"""
    return f"{h:.2f} hrs"


def _status_label(status_val: str, ot: float) -> str:
    m = {
        "present":    "Present",
        "present_ot": f"Present + {ot:.2f}h OT",
        "half_day":   "Half Day",
        "incomplete": "Incomplete",
        "absent":     "Absent",
        "lop":        "LOP",
    }
    return m.get(status_val, status_val.title())


# ── row builder ───────────────────────────────────────────────────────────── #

def _build_row(sno: int, user: User, rec: ProcessedAttendance) -> dict:
    sessions = []
    if rec.punch_sessions:
        try:
            sessions = json.loads(rec.punch_sessions)
        except Exception:
            pass

    hours = rec.work_duration_hours or 0.0
    ot    = rec.overtime_hours      or 0.0

    s1_in  = _fmt_time(sessions[0]["in"])  if len(sessions) > 0 else "-"
    s1_out = _fmt_time(sessions[0]["out"]) if len(sessions) > 0 else "-"
    s2_in  = _fmt_time(sessions[1]["in"])  if len(sessions) > 1 else "-"
    s2_out = _fmt_time(sessions[1]["out"]) if len(sessions) > 1 else "-"

    return {
        "sno":                  sno,
        "id":                   user.uid,
        "employee_name":        user.name,
        "date":                 rec.date.strftime("%d.%m.%Y"),
        "in1":                  s1_in,
        "out1":                 s1_out,
        "in2":                  s2_in,
        "out2":                 s2_out,
        "shift":                rec.shift or "Regular",
        "total_duration_hours": round(hours, 2),
        "total_duration_label": _fmt_hours(hours),
        "overtime_hours":       round(ot, 2),
        "status":               _status_label(rec.status.value if rec.status else "absent", ot),
        "remarks":              rec.remarks or "",
    }


def _month_records(db, uid: int, year: int, month: int):
    start = date(year, month, 1)
    end   = date(year, month, calendar.monthrange(year, month)[1])
    return (
        db.query(ProcessedAttendance)
        .filter(
            ProcessedAttendance.uid  == uid,
            ProcessedAttendance.date >= start,
            ProcessedAttendance.date <= end,
        )
        .order_by(ProcessedAttendance.date)
        .all()
    )


def _month_summary(records: list) -> dict:
    total_h = sum(r.work_duration_hours or 0 for r in records)
    total_ot = sum(r.overtime_hours     or 0 for r in records)
    return {
        "total_days":        len(records),
        "present":           sum(1 for r in records if r.status in (AttendanceStatus.PRESENT, AttendanceStatus.PRESENT_OT)),
        "half_day":          sum(1 for r in records if r.status == AttendanceStatus.HALF_DAY),
        "incomplete":        sum(1 for r in records if r.status == AttendanceStatus.INCOMPLETE),
        "absent":            sum(1 for r in records if r.status in (AttendanceStatus.ABSENT, AttendanceStatus.LOP)),
        "total_hours":       round(total_h,  2),
        "overtime_hours":    round(total_ot, 2),
        "avg_hours_per_day": round(total_h / len(records), 2) if records else 0,
    }


# ── Excel builder ─────────────────────────────────────────────────────────── #

REPORT_HEADERS = [
    "Sno", "ID", "Employee Name", "Date",
    "InTime-1", "OutTime-1", "InTime-2", "OutTime-2",
    "Shift", "Total Duration", "Status", "Remarks",
]

COL_WIDTHS = [5, 6, 22, 12, 12, 12, 12, 12, 14, 16, 20, 45]


def _rows_to_aoa(rows: list, emp_name: str, month_label: str) -> list:
    """Convert row dicts to array-of-arrays for openpyxl / xlsxwriter."""
    aoa = [
        [f"ATTENDANCE REPORT — {emp_name}"],
        [f"Period: {month_label}"],
        [],
        REPORT_HEADERS,
    ]
    for r in rows:
        aoa.append([
            r["sno"], r["id"], r["employee_name"], r["date"],
            r["in1"], r["out1"], r["in2"], r["out2"],
            r["shift"], r["total_duration_label"], r["status"], r["remarks"],
        ])
    return aoa


def _write_excel_single(
    rows: list,
    summary: dict,
    emp_name: str,
    month_label: str,
    sheet_title: str,
) -> bytes:
    """Build an xlsx bytes object with one sheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]

        thin  = Side(style="thin")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="1F3864")   # dark navy
        sub_fill = PatternFill("solid", fgColor="D9E1F2")   # light blue
        alt_fill = PatternFill("solid", fgColor="F2F2F2")   # light grey

        # Title rows
        ws.append([f"ATTENDANCE REPORT — {emp_name}"])
        ws.append([f"Period: {month_label}"])
        ws.append([])

        # Header row
        ws.append(REPORT_HEADERS)
        hdr_row = ws.max_row
        for col_idx, hdr in enumerate(REPORT_HEADERS, 1):
            cell = ws.cell(row=hdr_row, column=col_idx)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

        # Data rows
        status_colors = {
            "Present":    "C6EFCE",   # green
            "Half Day":   "FFEB9C",   # yellow
            "Incomplete": "FFC7CE",   # red
            "Absent":     "E0E0E0",   # grey
            "LOP":        "FF0000",   # bright red
        }

        for i, r in enumerate(rows):
            ws.append([
                r["sno"], r["id"], r["employee_name"], r["date"],
                r["in1"], r["out1"], r["in2"], r["out2"],
                r["shift"], r["total_duration_label"], r["status"], r["remarks"],
            ])
            row_num = ws.max_row
            # colour by status
            status_key = r["status"].split(" + ")[0].strip()   # strip OT part
            fill_hex = status_colors.get(status_key, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_hex)

            for col_idx in range(1, len(REPORT_HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = bdr
                cell.alignment = Alignment(vertical="center", horizontal="center" if col_idx not in (3, 12) else "left")
                if col_idx == 11:          # Status column
                    cell.fill = row_fill
                elif i % 2 == 1:
                    cell.fill = alt_fill

        # Summary section
        ws.append([])
        ws.append(["MONTHLY SUMMARY"])
        for key, val in [
            ("Total Days Worked", summary["total_days"]),
            ("Present",           summary["present"]),
            ("Half Day",          summary["half_day"]),
            ("Incomplete",        summary["incomplete"]),
            ("Absent / LOP",      summary["absent"]),
            ("Total Hours",       f"{summary['total_hours']} h"),
            ("Overtime Hours",    f"{summary['overtime_hours']} h"),
            ("Avg Hours / Day",   f"{summary['avg_hours_per_day']} h"),
        ]:
            ws.append([key, val])

        # Freeze header row
        ws.freeze_panes = f"A{hdr_row + 1}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)   # rewind so StreamingResponse reads from the start
        return buf

    except ImportError:
        raise


def _write_excel_multi(months_data: list) -> bytes:
    """
    months_data: [{"sheet_title", "rows", "summary", "emp_name", "month_label"}, ...]
    Produces one sheet per month.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    thin    = Side(style="thin")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    status_colors = {
        "Present":    "C6EFCE",
        "Half Day":   "FFEB9C",
        "Incomplete": "FFC7CE",
        "Absent":     "E0E0E0",
        "LOP":        "FF0000",
    }

    for md in months_data:
        ws = wb.create_sheet(title=md["sheet_title"][:31])
        rows      = md["rows"]
        summary   = md["summary"]
        emp_name  = md["emp_name"]
        month_lbl = md["month_label"]

        ws.append([f"ATTENDANCE REPORT — {emp_name}"])
        ws.append([f"Period: {month_lbl}"])
        ws.append([])
        ws.append(REPORT_HEADERS)

        hdr_row = ws.max_row
        for col_idx in range(1, len(REPORT_HEADERS) + 1):
            cell = ws.cell(row=hdr_row, column=col_idx)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

        for i, r in enumerate(rows):
            ws.append([
                r["sno"], r["id"], r["employee_name"], r["date"],
                r["in1"], r["out1"], r["in2"], r["out2"],
                r["shift"], r["total_duration_label"], r["status"], r["remarks"],
            ])
            row_num  = ws.max_row
            st_key   = r["status"].split(" + ")[0].strip()
            row_fill = PatternFill("solid", fgColor=status_colors.get(st_key, "FFFFFF"))

            for col_idx in range(1, len(REPORT_HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = bdr
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col_idx not in (3, 12) else "left"
                )
                if col_idx == 11:
                    cell.fill = row_fill
                elif i % 2 == 1:
                    cell.fill = alt_fill

        ws.append([])
        ws.append(["MONTHLY SUMMARY"])
        for key, val in [
            ("Total Days Worked", summary["total_days"]),
            ("Present",           summary["present"]),
            ("Half Day",          summary["half_day"]),
            ("Incomplete",        summary["incomplete"]),
            ("Absent / LOP",      summary["absent"]),
            ("Total Hours",       f"{summary['total_hours']} h"),
            ("Overtime Hours",    f"{summary['overtime_hours']} h"),
            ("Avg Hours / Day",   f"{summary['avg_hours_per_day']} h"),
        ]:
            ws.append([key, val])

        ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)   # rewind so StreamingResponse reads from the start
    return buf


# ── CSV helper ────────────────────────────────────────────────────────────── #

def _write_csv(rows: list) -> str:
    import csv
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(REPORT_HEADERS)
    for r in rows:
        w.writerow([
            r["sno"], r["id"], r["employee_name"], r["date"],
            r["in1"], r["out1"], r["in2"], r["out2"],
            r["shift"], r["total_duration_label"], r["status"], r["remarks"],
        ])
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════ #
#  Routes                                                                      #
# ═══════════════════════════════════════════════════════════════════════════ #

@router.get("/monthly-report/{uid}")
async def monthly_report(
    uid:    int,
    year:   int = Query(...),
    month:  int = Query(..., ge=1, le=12),
    export: Optional[str] = Query(None, description="csv | excel"),
    db: Session = Depends(get_db),
):
    """Single-employee monthly report."""
    try:
        user = db.query(User).filter(User.uid == uid, User.is_active == True).first()
        if not user:
            return error_response(f"User UID {uid} not found", {"uid": uid})

        records    = _month_records(db, uid, year, month)
        rows       = [_build_row(i + 1, user, rec) for i, rec in enumerate(records)]
        summary    = _month_summary(records)
        month_name = calendar.month_name[month]
        month_lbl  = f"{month_name} {year}"

        if export == "csv":
            content  = _write_csv(rows)
            filename = f"Attendance_{user.name.replace(' ','_')}_{month_name}_{year}.csv"
            return StreamingResponse(
                iter([content]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        if export == "excel":
            xlsx = _write_excel_single(
                rows, summary, user.name, month_lbl,
                sheet_title=month_name[:31],
            )
            filename = f"Attendance_{user.name.replace(' ','_')}_{month_name}_{year}.xlsx"
            return StreamingResponse(
                xlsx,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        return success_response(
            f"Monthly report — {user.name} — {month_lbl}",
            {
                "employee": {"uid": user.uid, "name": user.name},
                "period":   {"year": year, "month": month, "month_name": month_name},
                "summary":  summary,
                "rows":     rows,
            },
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return error_response("Failed to generate report", {"error": str(e)})


@router.get("/monthly-report-all")
async def monthly_report_all(
    year:   int = Query(...),
    month:  int = Query(..., ge=1, le=12),
    export: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """All active employees — monthly report."""
    try:
        users      = db.query(User).filter(User.is_active == True).order_by(User.name).all()
        month_name = calendar.month_name[month]
        month_lbl  = f"{month_name} {year}"

        all_emp = []
        for user in users:
            records = _month_records(db, user.uid, year, month)
            rows    = [_build_row(i + 1, user, rec) for i, rec in enumerate(records)]
            summary = _month_summary(records)
            all_emp.append({
                "employee": {"uid": user.uid, "name": user.name},
                "summary":  summary,
                "rows":     rows,
            })

        if export == "csv":
            sno = 1
            all_rows = []
            for emp in all_emp:
                for r in emp["rows"]:
                    r["sno"] = sno
                    sno += 1
                    all_rows.append(r)
            content  = _write_csv(all_rows)
            filename = f"All_Employees_{month_name}_{year}.csv"
            return StreamingResponse(
                iter([content]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        if export == "excel":
            months_data = [
                {
                    "sheet_title": emp["employee"]["name"][:31],
                    "rows":        emp["rows"],
                    "summary":     emp["summary"],
                    "emp_name":    emp["employee"]["name"],
                    "month_label": month_lbl,
                }
                for emp in all_emp
            ]
            xlsx     = _write_excel_multi(months_data)
            filename = f"All_Employees_{month_name}_{year}.xlsx"
            return StreamingResponse(
                xlsx,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        return success_response(
            f"All-employees report — {month_lbl}",
            {
                "period":           {"year": year, "month": month, "month_name": month_name},
                "total_employees":  len(users),
                "employees":        all_emp,
            },
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return error_response("Failed", {"error": str(e)})


@router.get("/range-report/{uid}")
async def range_report(
    uid:        int,
    start_date: date = Query(...),
    end_date:   date = Query(...),
    export:     Optional[str] = Query(None, description="excel | csv"),
    db: Session = Depends(get_db),
):
    """
    Multi-month range report for a single employee.
    Excel export → one sheet per calendar month.
    """
    try:
        user = db.query(User).filter(User.uid == uid, User.is_active == True).first()
        if not user:
            return error_response(f"User UID {uid} not found", {"uid": uid})

        # Enumerate all (year, month) pairs in the range
        ym_pairs = []
        cur = start_date.replace(day=1)
        while cur <= end_date:
            ym_pairs.append((cur.year, cur.month))
            # advance to next month
            if cur.month == 12:
                cur = cur.replace(year=cur.year + 1, month=1)
            else:
                cur = cur.replace(month=cur.month + 1)

        months_data = []
        all_rows_csv = []
        sno = 1

        for yr, mo in ym_pairs:
            records    = _month_records(db, uid, yr, mo)
            rows       = [_build_row(sno + i, user, rec) for i, rec in enumerate(records)]
            summary    = _month_summary(records)
            month_name = calendar.month_name[mo]
            month_lbl  = f"{month_name} {yr}"
            sno       += len(rows)

            months_data.append({
                "sheet_title": f"{month_name[:10]} {yr}",
                "rows":        rows,
                "summary":     summary,
                "emp_name":    user.name,
                "month_label": month_lbl,
            })
            all_rows_csv.extend(rows)

        if export == "csv":
            content  = _write_csv(all_rows_csv)
            filename = f"Attendance_{user.name.replace(' ','_')}_{start_date}_{end_date}.csv"
            return StreamingResponse(
                iter([content]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        if export == "excel":
            xlsx     = _write_excel_multi(months_data)
            filename = f"Attendance_{user.name.replace(' ','_')}_{start_date}_{end_date}.xlsx"
            return StreamingResponse(
                xlsx,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        # JSON response
        return success_response(
            f"Range report — {user.name} — {start_date} to {end_date}",
            {
                "employee": {"uid": user.uid, "name": user.name},
                "period":   {"start": str(start_date), "end": str(end_date)},
                "months":   [
                    {
                        "month_label": md["month_label"],
                        "summary":     md["summary"],
                        "rows":        md["rows"],
                    }
                    for md in months_data
                ],
            },
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return error_response("Failed", {"error": str(e)})


# ── Legacy compat ─────────────────────────────────────────────────────────── #

@router.get("/detailed-report/{uid}")
async def detailed_report_legacy(
    uid:        int,
    start_date: date = Query(...),
    end_date:   date = Query(...),
    db: Session = Depends(get_db),
):
    """Legacy endpoint kept for UI backward-compatibility."""
    user = db.query(User).filter(User.uid == uid).first()
    if not user:
        return error_response(f"User UID {uid} not found", {"uid": uid})

    records     = (
        db.query(ProcessedAttendance)
        .filter(ProcessedAttendance.uid == uid,
                ProcessedAttendance.date >= start_date,
                ProcessedAttendance.date <= end_date)
        .order_by(ProcessedAttendance.date)
        .all()
    )
    total_days  = len(records)
    total_hours = sum(r.work_duration_hours or 0 for r in records)
    total_ot    = sum(r.overtime_hours      or 0 for r in records)

    return success_response("Detailed report", {
        "user":              {"uid": user.uid, "name": user.name, "card_no": user.card_no},
        "period":            {"start_date": str(start_date), "end_date": str(end_date),
                              "total_days": (end_date - start_date).days + 1},
        "summary":           {
            "total_worked_days":    total_days,
            "present_days":         sum(1 for r in records if r.status in (AttendanceStatus.PRESENT, AttendanceStatus.PRESENT_OT)),
            "half_days":            sum(1 for r in records if r.status == AttendanceStatus.HALF_DAY),
            "incomplete_days":      sum(1 for r in records if r.status == AttendanceStatus.INCOMPLETE),
            "leaves":               0, "late_days": 0, "early_leave_days": 0,
            "total_hours_worked":   round(total_hours, 2),
            "overtime_hours":       round(total_ot,    2),
            "average_hours_per_day":round(total_hours / total_days, 2) if total_days else 0,
            "attendance_rate":      0,
        },
        "monthly_breakdown": [], "weekly_breakdown": [],
        "shift_analysis":    {}, "leave_summary":    {},
    })