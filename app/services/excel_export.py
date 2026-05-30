# app/services/excel_export.py
"""
Excel Export Service - MS Office Compatible (.xlsx)
────────────────────────────────────────────────────────────────────────────────

Exports attendance and payroll data in Microsoft Excel format (.xlsx)
Compatible with MS Office, Google Sheets, LibreOffice
Uses openpyxl for better MS Office compatibility
"""

from datetime import date, datetime
from typing import List, Dict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.user import User
from app.models.attendance import ProcessedAttendance


class ExcelExportService:
    """Service for exporting data to Excel (.xlsx) format."""

    def __init__(self, db: Session):
        self.db = db

    def _style_header(self, ws, row_num: int, columns: int):
        """Apply professional header styling."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, columns + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def _style_data_rows(self, ws, start_row: int, end_row: int, columns: int):
        """Apply data row styling with alternating colors."""
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Alternating row colors
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row in range(start_row, end_row + 1):
            fill = gray_fill if row % 2 == 0 else white_fill
            for col in range(1, columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                cell.fill = fill

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_today_attendance(self) -> BytesIO:
        """
        Export today's attendance to Excel.

        Format:
        - S.No | Employee Name | First IN | Last OUT | Work Hours | Shift Type | Status
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Today Attendance"

        # Title
        ws['A1'] = f"Attendance Report - {date.today().strftime('%d-%b-%Y')}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["S.No", "Employee Name", "First IN", "Last OUT", "Work Hours", "Shift Type", "Status"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header)

        self._style_header(ws, 3, len(headers))

        # Fetch today's attendance
        today = date.today()
        attendance_records = (
            self.db.query(ProcessedAttendance, User)
            .join(User, ProcessedAttendance.uid == User.uid)
            .filter(ProcessedAttendance.date == today)
            .order_by(User.name)
            .all()
        )

        # Data rows
        row = 4
        for idx, (att, user) in enumerate(attendance_records, start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=user.name)
            ws.cell(row=row, column=3, value=att.first_in.strftime('%I:%M %p') if att.first_in else "-")
            ws.cell(row=row, column=4, value=att.last_out.strftime('%I:%M %p') if att.last_out else "-")
            ws.cell(row=row, column=5, value=f"{att.work_duration_hours:.2f}" if att.work_duration_hours else "0.00")
            ws.cell(row=row, column=6, value=att.shift or "Regular")
            ws.cell(row=row, column=7, value=att.status.value.upper() if att.status else "INCOMPLETE")
            row += 1

        # Style data rows
        if row > 4:
            self._style_data_rows(ws, 4, row - 1, len(headers))

        # Auto-adjust columns
        self._auto_adjust_columns(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_payroll_report(self, start_date: date, end_date: date) -> BytesIO:
        """
        Export payroll report to Excel (MS Office compatible).

        Format:
        - S.No | Employee Name | Days Present | Days Half-Day | Days Incomplete | Total Work Hours | Regular Days | Break Shift Days
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Report"

        # Title
        ws['A1'] = f"Payroll Report: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "S.No",
            "Employee Name",
            "Days Present",
            "Days Half-Day",
            "Days Incomplete",
            "Total Work Hours",
            "Regular Days",
            "Break Shift Days"
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header)

        self._style_header(ws, 3, len(headers))

        # Fetch all active users
        users = self.db.query(User).filter(User.is_active == True).order_by(User.name).all()

        # Data rows
        row = 4
        for idx, user in enumerate(users, start=1):
            # Fetch attendance for date range
            attendance_records = (
                self.db.query(ProcessedAttendance)
                .filter(
                    ProcessedAttendance.uid == user.uid,
                    ProcessedAttendance.date >= start_date,
                    ProcessedAttendance.date <= end_date
                )
                .all()
            )

            if not attendance_records:
                continue

            # Calculate statistics
            days_present = sum(1 for a in attendance_records if a.status.value in ['present', 'present_ot'])
            days_half_day = sum(1 for a in attendance_records if a.status.value == 'half_day')
            days_incomplete = sum(1 for a in attendance_records if a.status.value == 'incomplete')
            total_hours = sum(a.work_duration_hours or 0 for a in attendance_records)
            regular_days = sum(1 for a in attendance_records if a.shift == 'Regular')
            break_shift_days = sum(1 for a in attendance_records if a.shift == 'Break Shift')

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=user.name)
            ws.cell(row=row, column=3, value=days_present)
            ws.cell(row=row, column=4, value=days_half_day)
            ws.cell(row=row, column=5, value=days_incomplete)
            ws.cell(row=row, column=6, value=f"{total_hours:.2f}")
            ws.cell(row=row, column=7, value=regular_days)
            ws.cell(row=row, column=8, value=break_shift_days)
            row += 1

        # Style data rows
        if row > 4:
            self._style_data_rows(ws, 4, row - 1, len(headers))

        # Auto-adjust columns
        self._auto_adjust_columns(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_detailed_attendance(self, uid: int, start_date: date, end_date: date) -> BytesIO:
        """
        Export detailed attendance for a specific user.

        Format:
        - Date | Day | First IN | Last OUT | Work Hours | Shift Type | Status | Remarks
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Detailed Attendance"

        # Get user info
        user = self.db.query(User).filter(User.uid == uid).first()
        if not user:
            raise ValueError(f"User with UID {uid} not found")

        # Title
        ws['A1'] = f"Detailed Attendance Report - {user.name}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal="center")

        ws['A2'] = f"Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
        ws['A2'].font = Font(name='Calibri', size=11)
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Date", "Day", "First IN", "Last OUT", "Work Hours", "Shift Type", "Status", "Remarks"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=4, column=col, value=header)

        self._style_header(ws, 4, len(headers))

        # Fetch attendance records
        attendance_records = (
            self.db.query(ProcessedAttendance)
            .filter(
                ProcessedAttendance.uid == uid,
                ProcessedAttendance.date >= start_date,
                ProcessedAttendance.date <= end_date
            )
            .order_by(ProcessedAttendance.date)
            .all()
        )

        # Data rows
        row = 5
        for att in attendance_records:
            ws.cell(row=row, column=1, value=att.date.strftime('%d-%b-%Y'))
            ws.cell(row=row, column=2, value=att.date.strftime('%A'))
            ws.cell(row=row, column=3, value=att.first_in.strftime('%I:%M %p') if att.first_in else "-")
            ws.cell(row=row, column=4, value=att.last_out.strftime('%I:%M %p') if att.last_out else "-")
            ws.cell(row=row, column=5, value=f"{att.work_duration_hours:.2f}" if att.work_duration_hours else "0.00")
            ws.cell(row=row, column=6, value=att.shift or "Regular")
            ws.cell(row=row, column=7, value=att.status.value.upper() if att.status else "INCOMPLETE")
            ws.cell(row=row, column=8, value=att.remarks or "-")
            row += 1

        # Style data rows
        if row > 5:
            self._style_data_rows(ws, 5, row - 1, len(headers))

        # Summary section
        if attendance_records:
            total_hours = sum(a.work_duration_hours or 0 for a in attendance_records)
            days_present = sum(1 for a in attendance_records if a.status.value in ['present', 'present_ot'])
            days_half_day = sum(1 for a in attendance_records if a.status.value == 'half_day')

            summary_row = row + 2
            ws.cell(row=summary_row, column=1, value="SUMMARY")
            ws.cell(row=summary_row, column=1).font = Font(name='Calibri', size=11, bold=True)

            ws.cell(row=summary_row + 1, column=1, value="Total Days Present:")
            ws.cell(row=summary_row + 1, column=2, value=days_present)

            ws.cell(row=summary_row + 2, column=1, value="Total Days Half-Day:")
            ws.cell(row=summary_row + 2, column=2, value=days_half_day)

            ws.cell(row=summary_row + 3, column=1, value="Total Work Hours:")
            ws.cell(row=summary_row + 3, column=2, value=f"{total_hours:.2f}")

        # Auto-adjust columns
        self._auto_adjust_columns(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
